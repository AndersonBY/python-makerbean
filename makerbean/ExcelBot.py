# -*- coding: utf-8 -*-
# @Author: ander
# @Date:   2020-12-22 16:17:53
# @Last Modified by:   ander
# @Last Modified time: 2020-12-22 16:18:34
from openpyxl import Workbook, load_workbook
import csv


class ExcelBot(object):
    """docstring for ExcelBot"""

    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.添加数据 = self.add_data
        self.添加一行数据 = self.add_row
        self.提取一行数据 = self.get_row
        self.提取一列数据 = self.get_col
        self.保存文件 = self.save

    def add_data(self, data):
        try:
            for row in data:
                self.add_row(row)
        except Exception as e:
            print(e)

    def add_row(self, row):
        try:
            self.sheet.append(row)
        except Exception as e:
            print(e)

    def get_row(self, row):
        data = []
        if isinstance(row, int):
            for cell in self.sheet[row + 1]:
                data.append(cell.value)

        return data

    def get_col(self, col):
        data = []
        if isinstance(col, str):
            for cell in self.sheet[col]:
                data.append(cell.value)
        elif isinstance(col, int):
            # excel column counts from 1
            col += 1
            for row in self.sheet.iter_rows(min_col=col, max_col=col):
                data.append(row[0].value)

        return data

    def clear(self):
        self.workbook.remove_sheet(self.sheet)
        self.sheet = self.workbook.create_sheet("sheet1")

    def open(self, filename):
        if filename.endswith(".csv"):
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            with open(filename, encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    self.sheet.append(row)
        else:
            self.workbook = load_workbook(filename=filename)
            self.sheet = self.workbook.active

    def save(self, filename, ext="xlsx"):
        self.filename = filename
        if ext == "xlsx":
            self.workbook.save(filename=f"{filename}.xlsx")
        elif ext == "csv":
            with open(f"{filename}.csv", "w", newline="", encoding="utf-8") as f:
                c = csv.writer(f)
                for r in self.sheet.rows:
                    c.writerow([cell.value for cell in r])
